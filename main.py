import os
import time
import requests
from openpyxl import load_workbook, Workbook
from procedury import parsing_synapsenet, parsing_rusprofile, tablica

imya_faila = tablica()
if imya_faila != 'Oshibka':
    wb = load_workbook(tablica(), data_only=True)
    ws = wb.active
    massiv_inn = []
    for row in ws.iter_rows(min_row=2, min_col = 1, max_col=2, values_only=True):
        print(row)
        adres = row[1]
        name = row[0]
        inn = parsing_synapsenet(name, adres)
        if inn == 'Ошибка':
            inn = parsing_rusprofile(name, adres)
        print(inn)
        massiv_inn.append(inn)

    i = 2
    for inn in massiv_inn:
        adres_cell = 'C' + str(i)
        i+=1
        ws[adres_cell] = inn

    wb.save("res.xlsx")
else:
    time.sleep(10)






        #pometit yacheyky











